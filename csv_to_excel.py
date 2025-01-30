import pandas as pd
import streamlit as st
import io
import uuid
from openpyxl import load_workbook

def transform_csv_to_excel_with_template(csv_data, template_file, selected_columns, column_mapping, selected_category, sheet_name):
    """
    Memasukkan data CSV ke dalam template Excel yang diunggah pengguna.
    """

    if csv_data is None or csv_data.size == 0:
        st.error("File CSV yang diunggah kosong! Silakan unggah file dengan data.")
        st.stop()

    try:
        csv_text = io.StringIO(csv_data.getvalue().decode("utf-8"))  # Konversi ke teks
        data = pd.read_csv(csv_text, encoding="utf-8")  # Default UTF-8
        if data.empty:
            st.error("File CSV tidak mengandung data!")
            st.stop()
    except pd.errors.EmptyDataError:
        st.error("File CSV kosong atau tidak valid!")
        st.stop()
    except UnicodeDecodeError:
        try:
            data = pd.read_csv(io.StringIO(csv_data.getvalue().decode("latin1")), encoding="latin1")
        except Exception:
            st.error("Format encoding file tidak didukung! Coba simpan ulang file dengan UTF-8.")
            st.stop()

    # Pastikan kolom yang dipilih ada dalam CSV
    missing_columns = [col for col in selected_columns if col not in data.columns]
    if missing_columns:
        st.error(f"Kolom berikut tidak ditemukan dalam CSV: {', '.join(missing_columns)}")
        st.stop()

    # Pilih hanya kolom yang dipilih pengguna
    csv_df = data[selected_columns]

    # Filter berdasarkan kategori yang dipilih
    if selected_category != "Semua" and "Segment_Category" in csv_df.columns:
        csv_df = csv_df[csv_df["Segment_Category"] == selected_category]

    # Ganti nama kolom sesuai pemetaan
    csv_df = csv_df.rename(columns=column_mapping)

    # Buka template Excel
    template_bytes = io.BytesIO(template_file.getvalue())  
    book = load_workbook(template_bytes)  

    # Pastikan sheet tersedia
    if sheet_name not in book.sheetnames:
        return None, f"Sheet '{sheet_name}' tidak ditemukan dalam template!"

    sheet = book[sheet_name]  

    # Masukkan data ke sheet (dimulai dari baris kedua agar tidak menimpa header)
    for r_idx, row in enumerate(csv_df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Simpan hasil ke memori (bukan disk)
    output = io.BytesIO()
    book.save(output)
    output.seek(0)

    return output, None

# ============================ STREAMLIT UI ============================ #

st.title("Ekstrak Data ke Template Excel")

# **Upload file CSV**
uploaded_csv = st.file_uploader("Upload file CSV", type=["csv"])

# **Upload Template Excel**
uploaded_template = st.file_uploader("Upload Template Excel", type=["xlsx"])

if uploaded_csv and uploaded_template:
    try:
        # Membaca data CSV
        data = pd.read_csv(uploaded_csv, encoding="utf-8")
    except UnicodeDecodeError:
        data = pd.read_csv(uploaded_csv, encoding="latin1")
    except Exception:
        st.error("Terjadi kesalahan saat membaca file CSV. Pastikan formatnya benar.")
        st.stop()

    # Menampilkan daftar kolom yang bisa dipilih
    all_columns = data.columns.tolist()

    # Pilih kolom yang ingin dimasukkan ke Excel
    selected_columns = st.multiselect("Pilih kolom yang ingin dimasukkan", all_columns, default=all_columns[:3])

    # Pemetaan nama kolom
    column_mapping = {col: col for col in selected_columns}

    # Pilihan kategori (jika 'Segment_Category' tersedia)
    if "Segment_Category" in data.columns:
        kategori_unik = ["Semua"] + sorted(data["Segment_Category"].dropna().unique().tolist())
        selected_category = st.selectbox("Pilih Kategori", kategori_unik)
    else:
        selected_category = "Semua"

    # Pilih sheet dalam template Excel
    book = load_workbook(io.BytesIO(uploaded_template.getvalue()))
    sheet_name = st.selectbox("Pilih Sheet untuk Menyimpan Data", book.sheetnames)

    if st.button("Ekspor ke Template Excel"):
        output_excel, error = transform_csv_to_excel_with_template(
            uploaded_csv, uploaded_template, selected_columns, column_mapping, selected_category, sheet_name
        )

        if error:
            st.error(error)
        else:
            unique_filename = f"export_{uuid.uuid4().hex}.xlsx"
            st.download_button(
                label="Unduh File",
                data=output_excel,
                file_name=unique_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
